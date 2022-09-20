"""
    @ author: 032002602陈俊龙
    @ Document description： 这个文件实现的功能是从文本中将疫情数据批量提取到Excel文件中并生成图表
"""
import re
import os
from openpyxl import Workbook
from creat_chart import myChart
import time
from Demo01.setting import *
province_name = ['福建', '河北', '河南', '湖南', '湖北', '广东', '北京', '上海', '天津', '重庆', '四川', '黑龙江', '内蒙古', '西藏', '宁夏', '浙江', '山东',
                 '安徽', '江西', '山西', '吉林', '江苏', '云南', '贵州', '陕西', '甘肃', '青海', '辽宁', '广西', '新疆', '海南']


# 格式化省份名称的实现
def verify_name(name_str: str) -> str:
    for name in province_name:
        if name in name_str:
            return name


# 用于检测文本中是否存在与本土病例有关的内容
def verify_local_info(content: str) -> str:
    if re.search(r"[^\u4e00-\u9fa5]本土病例.*。", content) is None:
        return 0
    else:
        info = (re.search(r"[^\u4e00-\u9fa5]本土病例.*。", content)).group()
        return info


# 用于解决出现类似：本土新增x例（均在福建）的问题
def format_name(content: str, area_name: str, diction: dict, num):
    # 如果提取出的地区名称不在省份名称列表中，则需要格式化
    if area_name not in province_name:
        area_name = verify_name(area_name)  # 将省份名称格式化
    if re.search(r'[0-9]+', content) is None or content[1] == '均' or content[1] == '在':
        diction[area_name] = num
    else:
        # 提取出省份的新增数量
        increase_num = re.search(r'[0-9]+', content).group()
        diction[area_name] = increase_num


# 获取本土病例相关数据
def get_Local_data(content: str) -> dict:
    area_num = {'本土病例': 0}
    information = verify_local_info(content)  # 提取于本土病例相关内容
    if information == 0:  # 如果没有与新增本土病例相关的内容，直接返回
        area_num = refine_dict(area_num)    # 将所有省份数据设为0
        return area_num
    all_increase = (re.search(r'\d+', information)).group()  # all_increase 为本土新增的病例数
    area_num['本土病例'] = all_increase     # 本土新增病例的值为 all_increase
    # ----------------- 提取省份数据-------------------------
    # 获取出每个省份新增人数的文本信息
    province_info = (re.search(r'（.*?）', information)).group()
    # 解决文本格式不一致的问题和各个省份数据提取（对字符串进行切割）,比如说xx省x例，其中xx市几例；xx省x例，其中xx市x例
    if '；' in province_info:
        increase_list = province_info.split('；')
    else:
        increase_list = province_info.split('，')
    # 遍历切割完的字符串（即包含各个省份数据的列表，对其进行正则提取）
    for area_increase in increase_list:
        # 提取出省份的相关信息
        area_name = re.search(r'[\u4e00-\u9fa5]+', area_increase).group()
        # 避免一些不规范文本的出现，将省份数据格式化
        format_name(area_increase, area_name, area_num, all_increase)
    area_num = refine_dict(area_num)    # 将无新增病例的省份数据初始化为0
    return area_num


# 获取新增无症状感染者数据
def get_asymptomatic_infected_data(content):
    asymptomatic_dict = {'新增无症状感染者': 0}  # 定义一个字典来存储无症状新增数据
    if re.search(r'本土[0-9]+例（.*?）', content) is None:   # 如果正则获取不到我们需要的文本信息，直接返回
        return asymptomatic_dict
    else:
        # 获取本土新增无症状的文本信息
        mainland_info = re.search(r'本土[0-9]+例（.*?）', content).group()
        # 获取本土新增无症状的数量
        asymptomatic_increase = re.search(r'[0-9]+', mainland_info).group()
        asymptomatic_dict['新增无症状感染者'] = asymptomatic_increase
        # 正则提取出与省份有关的文本信息
        province_asymptomatic_increase = re.search(r'（.*?）', mainland_info).group()
        # 对省份文本信息进行切割
        if len(province_asymptomatic_increase.split('；')) == 1:
            province_list = province_asymptomatic_increase.split('，')
        else:
            province_list = province_asymptomatic_increase.split('；')
        # 遍历切割后形成的列表
        for province in province_list:
            area_name = re.search(r'[\u4e00-\u9fa5]+', province).group()    # 提取出含省份数据的文本
            # 处理特殊格式，并格式化省份名称
            format_name(province, area_name, asymptomatic_dict, asymptomatic_increase)
        asymptomatic_dict = refine_dict(asymptomatic_dict)  # 将无新增无症状感染者的省份数据初始化为0
        return asymptomatic_dict


# 改进字典，将没有新增病例或无症状感染者的省份的值设为0
def refine_dict(dictionary):
    for province in province_name:
        # 如果字典的key中没有这个省份
        if province not in dictionary:
            # 将这个省份的数量置为0
            dictionary[province] = 0
    return dictionary


# 定义一个Excel写入函数，用于大量数据的处理。
def excel_write(worksheet, data):
    name = list(data.keys())
    value = list(data.values())
    long = len(name)
    # 开始写入Excel
    for row in range(1, long + 1):
        worksheet.cell(row=row, column=1, value=name[row-1])
        worksheet.cell(row=row, column=2, value=int(value[row-1]))


# 将数据保存至excel中
def save_to_excel(data_1: dict, data_2: dict, filename):
    wb = Workbook()  # 新建一个Workbook对象
    wb.remove(wb.active)    # 删除当前Workbook自带的sheet
    worksheet_1 = wb.create_sheet("新增本土病例")    # 在工作簿中新建一个sheet，用来保存本土病例数据
    worksheet_2 = wb.create_sheet("新增无症状感染者")   # 新建一个sheet，保存新增无症状数据
    # 以下是优化后的Excel写入函数，在进行大量文件处理时用得到。
    excel_write(worksheet_1, data_1)    # 在sheet中写入本土病例数据
    excel_write(worksheet_2, data_2)    # 在另一个sheet中写入数据新增无症状数据
    wb.save(f'{EXCEL_SAVE_PATH}/{filename}.xlsx')  # 保存为Excel
    myChart(f'{EXCEL_SAVE_PATH}/{filename}.xlsx')


# 这个函数用于提取当日最新数据
def draw_today_data(file):
    f = open(file=f'{TXT_SAVE_PATH}/{file}.txt', mode='r', encoding='utf-8')
    text = f.read()
    local_data = get_Local_data(text)
    asymptomatic_data = get_asymptomatic_infected_data(text)
    # 保存到Excel中
    save_to_excel(local_data, asymptomatic_data, file)
    f.close()
    print(f"{file}提取完成")


# 主函数，用于向其他文件暴露调用接口
def main():
    start = time.time()  # 设置时间戳，查看程序运行时间
    path = f'{TXT_SAVE_PATH}'
    dirs = os.listdir(path)  # 获取所有的文件
    dirs.reverse()  # 将文件顺序倒置，从最新的开始提取
    for file in dirs:
        if file == '2021-05-15-截至5月14日24时新型冠状病毒肺炎疫情最新情况.txt':
            break
        f = open(file=f'{TXT_SAVE_PATH}/{file}', mode='r', encoding='utf-8')
        text = f.read()
        local_data = get_Local_data(text)
        asymptomatic_data = get_asymptomatic_infected_data(text)
        # 将数据提取到Excel文件中
        save_to_excel(local_data, asymptomatic_data, file)
        f.close()
        print(f"{file}提取完成")
    end = time.time()
    print(f"所有文件提取完成，所用时间：{end - start}s")


if __name__ == '__main__':
    main()